from __future__ import annotations

import calendar
import base64
import csv
import html
import importlib.util
import math
import queue
import threading
import unicodedata
from datetime import date, datetime, time, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import colorchooser, filedialog, messagebox, simpledialog, ttk

from ayarlar import ConfigError, load_config, save_config
from ik_takip_hesaplama import (
    WorkRules,
    calculate_row,
    calculate_totals,
    format_hours,
    minutes_between,
    normalize_kind,
    parse_date,
    rules_from_config,
    sum_overlaps,
    working_periods,
)


def date_format_to_strftime(display_format: str) -> str:
    return (
        display_format
        .replace("yyyy", "%Y")
        .replace("YYYY", "%Y")
        .replace("dd", "%d")
        .replace("DD", "%d")
        .replace("MM", "%m")
    )


def normalize_search(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value.casefold())
    return "".join(char for char in normalized if not unicodedata.combining(char))


class CalendarPopup(tk.Toplevel):
    def __init__(self, master: tk.Tk, selected: date, on_select, theme: dict) -> None:
        super().__init__(master)
        self.selected = selected
        self.current_year = selected.year
        self.current_month = selected.month
        self.on_select = on_select
        self.theme = theme
        self.title("Tarih Seç")
        self.resizable(False, False)
        self.configure(bg=theme["background_color"])
        self.transient(master)
        self.grab_set()
        self.header = ttk.Frame(self)
        self.header.pack(fill="x", padx=8, pady=8)
        ttk.Button(self.header, text="<", command=self.previous_month, width=3).pack(side="left")
        self.title_var = tk.StringVar()
        ttk.Label(self.header, textvariable=self.title_var, width=18, anchor="center").pack(side="left", padx=6)
        ttk.Button(self.header, text=">", command=self.next_month, width=3).pack(side="left")
        self.days_frame = ttk.Frame(self)
        self.days_frame.pack(padx=8, pady=(0, 8))
        self.render()

    def previous_month(self) -> None:
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self.render()

    def next_month(self) -> None:
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self.render()

    def render(self) -> None:
        for widget in self.days_frame.winfo_children():
            widget.destroy()
        self.title_var.set(f"{self.current_month:02d}.{self.current_year}")
        day_names = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
        for index, name in enumerate(day_names):
            ttk.Label(self.days_frame, text=name, anchor="center").grid(row=0, column=index, padx=2, pady=2)
        month_days = calendar.monthcalendar(self.current_year, self.current_month)
        for row_index, week in enumerate(month_days, start=1):
            for column_index, day_number in enumerate(week):
                if day_number == 0:
                    ttk.Label(self.days_frame, text="", width=4).grid(row=row_index, column=column_index, padx=2, pady=2)
                    continue
                day_value = date(self.current_year, self.current_month, day_number)
                ttk.Button(
                    self.days_frame,
                    text=str(day_number),
                    width=4,
                    command=lambda value=day_value: self.select(value),
                ).grid(row=row_index, column=column_index, padx=2, pady=2)

    def select(self, value: date) -> None:
        self.on_select(value)
        self.destroy()


class ToolTip:
    def __init__(self, widget: tk.Widget, text: str) -> None:
        self.widget = widget
        self.text = text
        self.window: tk.Toplevel | None = None
        widget.bind("<Enter>", self.show)
        widget.bind("<Leave>", self.hide)

    def show(self, _event=None) -> None:
        if self.window or not self.text:
            return
        x = self.widget.winfo_rootx() + 18
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 8
        self.window = tk.Toplevel(self.widget)
        self.window.wm_overrideredirect(True)
        self.window.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            self.window,
            text=self.text,
            background="#FFFFE0",
            foreground="#2B2B2B",
            relief="solid",
            borderwidth=1,
            padx=8,
            pady=4,
            font=("Segoe UI", 9),
        )
        label.pack()

    def hide(self, _event=None) -> None:
        if self.window:
            self.window.destroy()
            self.window = None


CALCULATED_KEYS = {
    "Haftaiçi Saat": "haftaici_dk",
    "Cumartesi Saat": "cumartesi_dk",
    "Resmi Tatil Saat": "resmi_tatil_mesai_dk",
    "Toplam Saat": "toplam_dk",
}
PERSONNEL_RECORD_COLUMNS = [
    "Ä°sim",
    "Departman",
    "TÃ¼r",
    "BaÅŸlangÄ±Ã§ Tarihi",
    "BaÅŸlangÄ±Ã§ Saati",
    "BitiÅŸ Tarihi",
    "BitiÅŸ Saati",
    "HaftaiÃ§i Saat",
    "Cumartesi Saat",
    "Resmi Tatil Saat",
    "Toplam Saat",
    "AÃ§Ä±klama",
]
SETTINGS_PASSWORD = "1111"


class PersonelTakipApp(tk.Tk):
    def __init__(self, config: dict) -> None:
        super().__init__()
        self.config_data = config
        self.theme = config.get("theme", {})
        self.date_settings = config.get("date_settings", {})
        self.date_format = self.date_settings.get("display_format", "dd.MM.yyyy")
        self.date_strftime = date_format_to_strftime(self.date_format)
        self.rules = rules_from_config(config)
        self.report_missing_modules = self.missing_report_modules()
        self.columns = list(config["columns"])
        self.input_columns = list(config["input_columns"])
        self.personnel_record_columns = self.resolve_personnel_record_columns()
        self.rows: list[dict[str, str]] = []
        self.filtered_indices: list[int] = []
        self.current_page = 0
        performance = config.get("performance", {})
        self.page_size = int(performance.get("page_size", 1000))
        self.csv_progress_interval = int(performance.get("csv_progress_interval", 50000))
        self.csv_queue: queue.Queue = queue.Queue()
        self.kpi_queue: queue.Queue = queue.Queue()
        self.loading_csv = False
        self.filtering_rows = False
        self.filter_generation = 0
        self.kpi_running = False
        self.entries: dict[str, tk.Widget] = {}
        self.combo_widgets: dict[str, ttk.Combobox] = {}
        self.entry_vars: dict[str, tk.StringVar] = {}
        self.filter_vars: dict[str, tk.StringVar] = {}
        self.summary_vars = {
            "izin": tk.StringVar(value="0:00"),
            "mesai": tk.StringVar(value="0:00"),
            "rapor": tk.StringVar(value="0:00"),
            "resmi": tk.StringVar(value="0:00"),
        }
        self.status_var = tk.StringVar(value="")
        self.page_var = tk.StringVar(value="")
        self.kpi_person_var = tk.StringVar()
        self.kpi_person_count_var = tk.StringVar()
        self.kpi_month_var = tk.StringVar(value=str(date.today().month))
        self.kpi_year_var = tk.StringVar(value=str(date.today().year))
        self.kpi_status_var = tk.StringVar(value="")
        self.kpi_card_vars: dict[str, tk.StringVar] = {}

        program = config["program"]
        self.title(program["name"])
        self.geometry(f'{program.get("window_width", 1220)}x{program.get("window_height", 740)}')
        self.minsize(980, 620)
        self.apply_theme()
        self.create_widgets()
        self.set_default_dates()
        self.refresh_table()
        first_column = self.input_columns[0]
        if first_column in self.entries:
            self.entries[first_column].focus_set()

    def label(self, key: str, fallback: str) -> str:
        return self.config_data.get("labels", {}).get(key, fallback)

    def color(self, key: str, fallback: str | None = None) -> str:
        if key in self.theme:
            return self.theme[key]
        if fallback is not None:
            return fallback
        return self.theme["background_color"]

    def resolve_personnel_record_columns(self) -> list[str]:
        configured_columns = list(self.config_data.get("columns", []))
        if configured_columns:
            return configured_columns
        return list(PERSONNEL_RECORD_COLUMNS)

    def personnel_export_row(self, row: dict[str, str]) -> dict[str, str]:
        display = self.display_row(row)
        return {column: display.get(column, "") for column in self.personnel_record_columns}

    def personnel_export_values(self, row: dict[str, str]) -> list[str]:
        export_row = self.personnel_export_row(row)
        return [export_row.get(column, "") for column in self.personnel_record_columns]

    def is_date_column(self, column: str) -> bool:
        return "Tarihi" in column

    def format_date(self, value: date) -> str:
        return value.strftime(self.date_strftime)

    def missing_report_modules(self) -> list[str]:
        required = {
            "PIL": "PIL",
            "reportlab": "reportlab",
            "openpyxl": "openpyxl",
        }
        return [label for label, module_name in required.items() if importlib.util.find_spec(module_name) is None]

    def report_dependency_message(self, missing: list[str] | None = None) -> str:
        missing = missing if missing is not None else self.missing_report_modules()
        missing_text = ", ".join(missing) if missing else "-"
        return (
            "Rapor oluşturmak için gerekli kütüphaneler eksik.\n\n"
            f"Eksik modül: {missing_text}\n\n"
            "Kurulum:\n"
            "pip install pillow reportlab openpyxl"
        )

    def ensure_report_modules_available(self) -> bool:
        missing = self.missing_report_modules()
        self.report_missing_modules = missing
        if missing:
            messagebox.showerror("KPI Raporu", self.report_dependency_message(missing))
            return False
        return True

    def apply_theme(self) -> None:
        self.configure(bg=self.color("background_color"))
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure("TFrame", background=self.color("background_color"))
        style.configure("TLabel", background=self.color("background_color"), foreground=self.color("text_color"))
        style.configure("TLabelframe", background=self.color("background_color"), foreground=self.color("text_color"))
        style.configure("TLabelframe.Label", background=self.color("background_color"), foreground=self.color("text_color"))
        style.configure("Card.TFrame", background=self.color("card_background_color", "#FFFDFD"))
        style.configure("Card.TLabel", background=self.color("card_background_color", "#FFFDFD"), foreground=self.color("text_color"))
        style.configure("TNotebook", background=self.color("background_color"), borderwidth=0)
        style.configure(
            "TNotebook.Tab",
            background=self.color("secondary_color"),
            foreground=self.color("text_color"),
            padding=(12, 6),
        )
        style.map(
            "TNotebook.Tab",
            background=[("selected", self.color("primary_color"))],
            foreground=[("selected", self.color("button_text_color"))],
        )
        style.configure(
            "TButton",
            background=self.color("button_color"),
            foreground=self.color("button_text_color"),
            bordercolor=self.color("accent_color", self.color("primary_color")),
            focusthickness=1,
            focuscolor=self.color("accent_color", self.color("primary_color")),
            padding=(10, 6),
        )
        style.map(
            "TButton",
            background=[("active", self.color("secondary_color")), ("pressed", self.color("primary_color"))],
            foreground=[("active", self.color("button_text_color"))],
        )
        style.configure(
            "TEntry",
            fieldbackground=self.color("input_background_color"),
            foreground=self.color("text_color"),
            bordercolor=self.color("primary_color"),
            insertcolor=self.color("text_color"),
        )
        style.configure(
            "TCombobox",
            fieldbackground=self.color("input_background_color"),
            foreground=self.color("text_color"),
            bordercolor=self.color("primary_color"),
            arrowcolor=self.color("text_color"),
        )
        style.configure(
            "Treeview",
            background=self.color("table_background_color"),
            fieldbackground=self.color("table_background_color"),
            foreground=self.color("text_color"),
            bordercolor=self.color("primary_color"),
            rowheight=24,
        )
        style.configure(
            "Treeview.Heading",
            background=self.color("secondary_color"),
            foreground=self.color("text_color"),
        )
        style.map("Treeview", background=[("selected", self.color("primary_color"))])

    def create_widgets(self) -> None:
        header = ttk.Frame(self)
        header.pack(fill="x", padx=12, pady=(10, 0))
        self.create_logo(header)

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=12, pady=12)

        tabs = self.config_data.get("tabs", ["Personel Kayıtları", "Personel KPI", "Personel Ayarları", "Genel Ayarlar"])
        main_tab = ttk.Frame(notebook)
        notebook.add(main_tab, text=tabs[0])
        self.create_records_tab(main_tab)

        kpi_tab = ttk.Frame(notebook)
        notebook.add(kpi_tab, text=tabs[1] if len(tabs) > 1 else "Personel KPI")
        self.create_kpi_tab(kpi_tab)

        personnel_settings_tab = ttk.Frame(notebook)
        notebook.add(personnel_settings_tab, text=tabs[2] if len(tabs) > 2 else "Personel Ayarları")
        self.create_personnel_settings_tab(personnel_settings_tab)

        general_settings_tab = ttk.Frame(notebook)
        notebook.add(general_settings_tab, text=tabs[3] if len(tabs) > 3 else "Genel Ayarlar")
        self.create_general_settings_tab(general_settings_tab)

    def create_logo(self, parent: ttk.Frame) -> None:
        program = self.config_data["program"]
        logo_config = self.config_data["logo"]
        font = (logo_config["font"], 28, "bold")
        logo = ttk.Frame(parent)
        logo.pack(side="left")
        text = logo_config.get("text", "HS")
        first = text[:1] or "H"
        rest = text[1:] or "S"
        self.logo_first_label = tk.Label(logo, text=first, font=font, fg=logo_config["h_color"], bg=self.color("background_color"))
        self.logo_first_label.pack(side="left")
        self.logo_rest_label = tk.Label(logo, text=rest, font=font, fg=logo_config["s_color"], bg=self.color("background_color"))
        self.logo_rest_label.pack(side="left")
        self.program_title_label = ttk.Label(parent, text=program["name"], font=("", 14, "bold"))
        self.program_title_label.pack(side="left", padx=12)

    def update_logo_display(self) -> None:
        logo_config = self.config_data.get("logo", {})
        text = logo_config.get("text", "HS")
        font = (logo_config.get("font", "Segoe UI Semibold"), 28, "bold")
        if hasattr(self, "logo_first_label"):
            self.logo_first_label.configure(
                text=text[:1] or "H",
                font=font,
                fg=logo_config.get("h_color", "#001F5B"),
                bg=self.color("background_color"),
            )
        if hasattr(self, "logo_rest_label"):
            self.logo_rest_label.configure(
                text=text[1:] or "S",
                font=font,
                fg=logo_config.get("s_color", "#000000"),
                bg=self.color("background_color"),
            )
        if hasattr(self, "program_title_label"):
            self.program_title_label.configure(text=self.config_data["program"]["name"])

    def create_records_tab(self, parent: ttk.Frame) -> None:
        form = ttk.LabelFrame(parent, text=self.label("new_record", "Yeni kayıt"))
        form.pack(fill="x", padx=4, pady=(4, 10))

        for index, column in enumerate(self.input_columns):
            ttk.Label(form, text=column).grid(row=0, column=index, sticky="w", padx=5, pady=(8, 2))
            variable = tk.StringVar(value=self.config_data.get("defaults", {}).get(column, ""))
            self.entry_vars[column] = variable
            options = self.config_data.get("select_options", {}).get(column)
            if options is not None:
                widget = ttk.Combobox(form, textvariable=variable, values=options, width=15)
                self.combo_widgets[column] = widget
                if column == "İsim":
                    self.attach_person_autocomplete(widget, variable)
            else:
                widget = ttk.Entry(form, textvariable=variable, width=15)
            if self.is_date_column(column) and self.date_settings.get("use_calendar_picker", True):
                widget.bind("<Button-1>", lambda _event, field=column: self.open_calendar(field))
            widget.grid(row=1, column=index, sticky="ew", padx=5, pady=(0, 8))
            self.entries[column] = widget
            form.columnconfigure(index, weight=1)

        button_frame = ttk.Frame(form)
        button_frame.grid(row=1, column=len(self.input_columns), sticky="e", padx=6, pady=(0, 8))
        ttk.Button(button_frame, text=self.label("add", "Ekle"), command=self.add_row).pack(side="left", padx=3)
        ttk.Button(button_frame, text=self.label("delete_selected", "Seçileni Sil"), command=self.delete_selected).pack(side="left", padx=3)

        toolbar = ttk.Frame(parent)
        toolbar.pack(fill="x", padx=4, pady=(0, 8))
        ttk.Button(toolbar, text=self.label("open_csv", "CSV Aç"), command=self.open_csv).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text=self.label("save_csv", "CSV Kaydet"), command=self.save_csv).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text=self.label("save_excel", "Excel Kaydet"), command=self.save_excel).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text=self.label("example_row", "Örnek Satır"), command=self.fill_example).pack(side="left")
        ttk.Button(toolbar, text=self.label("clear_filters", "Filtreleri Temizle"), command=self.clear_filters).pack(side="left", padx=(6, 0))

        filter_frame = ttk.Frame(parent)
        filter_frame.pack(fill="x", padx=4, pady=(0, 4))
        search_fields = self.config_data.get("search_fields", self.columns)
        for index, column in enumerate(self.columns):
            filter_var = tk.StringVar()
            filter_var.trace_add("write", lambda *_args: self.on_filter_changed())
            self.filter_vars[column] = filter_var
            state = "normal" if column in search_fields else "disabled"
            if column == "İsim":
                entry = ttk.Combobox(filter_frame, textvariable=filter_var, values=self.person_options(include_everyone=True), width=12, state=state)
                self.filter_person_combo = entry
                self.attach_person_autocomplete(entry, filter_var, include_everyone=True)
            else:
                entry = ttk.Entry(filter_frame, textvariable=filter_var, width=12, state=state)
            entry.grid(row=0, column=index, sticky="ew", padx=1)
            filter_frame.columnconfigure(index, weight=1)

        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill="both", expand=True, padx=4)
        self.tree = ttk.Treeview(tree_frame, columns=self.columns, show="headings", selectmode="extended")
        tree_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        for column in self.columns:
            self.tree.heading(column, text=column)
            self.tree.column(column, width=120, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")
        self.tree.bind("<<TreeviewSelect>>", self.load_selected_into_form)
        self.tree.bind("<Control-c>", self.copy_selected_rows)
        self.tree.bind("<Control-C>", self.copy_selected_rows)
        self.tree.bind("<Control-v>", self.paste_rows)
        self.tree.bind("<Control-V>", self.paste_rows)
        self.tree.tag_configure("odd", background=self.color("table_background_color"))
        self.tree.tag_configure("even", background=self.color("zebra_color", "#FBE3E5"))

        pager = ttk.Frame(parent)
        pager.pack(fill="x", padx=4, pady=(6, 0))
        ttk.Button(pager, text="Önceki Sayfa", command=self.previous_page).pack(side="left", padx=(0, 6))
        ttk.Button(pager, text="Sonraki Sayfa", command=self.next_page).pack(side="left", padx=(0, 12))
        ttk.Label(pager, textvariable=self.page_var).pack(side="left", padx=(0, 12))
        ttk.Label(pager, textvariable=self.status_var).pack(side="left", padx=(0, 12))
        self.progress = ttk.Progressbar(pager, mode="indeterminate", length=160)
        self.progress.pack(side="right")

        self.create_summary(parent)

    def create_kpi_tab(self, parent: ttk.Frame) -> None:
        filters = ttk.LabelFrame(parent, text="Filtreler")
        filters.pack(fill="x", padx=8, pady=8)

        ttk.Label(filters, text="Personel").grid(row=0, column=0, sticky="w", padx=6, pady=(6, 2))
        self.kpi_person_var.set("Herkes")
        self.kpi_person_count_var.set(str(self.default_kpi_person_count()))
        self.kpi_person_combo = ttk.Combobox(filters, textvariable=self.kpi_person_var, values=self.person_options(include_everyone=True), width=24)
        self.kpi_person_combo.grid(row=1, column=0, sticky="ew", padx=6, pady=(0, 8))
        self.attach_person_autocomplete(self.kpi_person_combo, self.kpi_person_var, include_everyone=True)

        ttk.Label(filters, text="Ay").grid(row=0, column=1, sticky="w", padx=6, pady=(6, 2))
        month_values = [str(index) for index in range(1, 13)]
        ttk.Combobox(filters, textvariable=self.kpi_month_var, values=month_values, width=8).grid(row=1, column=1, sticky="ew", padx=6, pady=(0, 8))

        ttk.Label(filters, text="Yıl").grid(row=0, column=2, sticky="w", padx=6, pady=(6, 2))
        years = [str(date.today().year + offset) for offset in range(-3, 4)]
        ttk.Combobox(filters, textvariable=self.kpi_year_var, values=years, width=10).grid(row=1, column=2, sticky="ew", padx=6, pady=(0, 8))
        ttk.Label(filters, text="Personel Sayısı").grid(row=0, column=3, sticky="w", padx=6, pady=(6, 2))
        ttk.Entry(filters, textvariable=self.kpi_person_count_var, width=10).grid(row=1, column=3, sticky="ew", padx=6, pady=(0, 8))
        ttk.Button(filters, text="Yenile", command=self.refresh_kpi).grid(row=1, column=4, sticky="w", padx=6, pady=(0, 8))
        self.kpi_report_button = ttk.Button(filters, text="📄 KPI Raporu Oluştur", command=self.create_kpi_report)
        self.kpi_report_button.grid(row=1, column=5, sticky="w", padx=6, pady=(0, 8))
        if self.report_missing_modules:
            self.kpi_report_button.configure(state="disabled")
            ToolTip(self.kpi_report_button, "Rapor sistemi için ek bileşenler gerekli.")
        ttk.Label(filters, textvariable=self.kpi_status_var).grid(row=1, column=6, sticky="w", padx=6, pady=(0, 8))
        for column in range(7):
            filters.columnconfigure(column, weight=1)

        cards = ttk.Frame(parent)
        cards.pack(fill="x", padx=8, pady=(0, 8))
        card_titles = [
            ("fazla_mesai", "Fazla Mesai Saati"),
            ("izin", "Kullanılan İzin Saati"),
            ("rapor", "Rapor Saati"),
            ("resmi", "Resmi Tatil Mesai Saati"),
            ("toplam_kayip", "Toplam İşçilik Kaybı"),
            ("toplam_kapasite", "Toplam Kapasite"),
            ("kullanilabilir", "Eksik Çalışmalar Hariç Kapasite"),
            ("ek_kapasite", "Fazla Mesai ile Kazanılan Ek Kapasite"),
            ("eksik_oran", "Eksik Çalışma Oranı %"),
            ("kalan_limit", "Kişi Bazlı Kalan Mesai Limiti"),
        ]
        for index, (key, title) in enumerate(card_titles):
            frame = tk.Frame(
                cards,
                bg=self.color("card_background_color", "#FFFDFD"),
                bd=0,
                relief="flat",
                highlightthickness=1,
                highlightbackground=self.color("secondary_color"),
            )
            row = index // 5
            column = index % 5
            frame.grid(row=row, column=column, sticky="ew", padx=6, pady=6)
            tk.Frame(frame, bg=self.color("accent_color", "#001F5B"), height=3).pack(fill="x")
            tk.Label(
                frame,
                text=title,
                bg=self.color("card_background_color", "#FFFDFD"),
                fg=self.color("text_color"),
                font=("Segoe UI", 9, "bold"),
            ).pack(anchor="w", padx=10, pady=(8, 2))
            var = tk.StringVar(value="%0.00" if key == "eksik_oran" else "0:00 / 0 gün")
            self.kpi_card_vars[key] = var
            tk.Label(
                frame,
                textvariable=var,
                bg=self.color("card_background_color", "#FFFDFD"),
                fg=self.color("accent_color", "#001F5B"),
                font=("Segoe UI", 15, "bold"),
            ).pack(anchor="w", padx=10, pady=(0, 8))
        for column in range(5):
            cards.columnconfigure(column, weight=1)

        charts = ttk.Frame(parent)
        charts.pack(fill="both", expand=True, padx=8, pady=4)
        self.loss_canvas = tk.Canvas(charts, height=230, bg=self.color("chart_background_color", "#FFFFFF"), highlightthickness=1, highlightbackground=self.color("secondary_color"))
        self.extra_capacity_canvas = tk.Canvas(charts, height=230, bg=self.color("chart_background_color", "#FFFFFF"), highlightthickness=1, highlightbackground=self.color("secondary_color"))
        self.loss_canvas.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)
        self.extra_capacity_canvas.grid(row=0, column=1, sticky="nsew", padx=6, pady=6)
        charts.columnconfigure(0, weight=1)
        charts.columnconfigure(1, weight=1)
        charts.rowconfigure(0, weight=1)

        top_frame = ttk.LabelFrame(parent, text="Top Mesai Listesi")
        top_frame.pack(fill="both", expand=False, padx=8, pady=(4, 8))
        top_columns = ["Personel", "Fazla Mesai Saati", "Resmi Tatil Mesai Saati", "Toplam Mesai", "Mesai Limit Kullanım Oranı %"]
        self.top_overtime_tree = ttk.Treeview(top_frame, columns=top_columns, show="headings", height=7)
        for column in top_columns:
            self.top_overtime_tree.heading(column, text=column)
            self.top_overtime_tree.column(column, width=160, anchor="center")
        self.top_overtime_tree.tag_configure("odd", background=self.color("table_background_color"))
        self.top_overtime_tree.tag_configure("even", background=self.color("zebra_color", "#FBE3E5"))
        self.top_overtime_tree.pack(fill="both", expand=True, padx=6, pady=6)

    def default_kpi_person_count(self) -> int:
        configured = self.config_data.get("select_options", {}).get("İsim", [])
        configured_count = len({item.strip() for item in configured if item.strip()})
        if configured_count:
            return configured_count
        return len({row.get("İsim", "").strip() for row in self.rows if row.get("İsim", "").strip()})

    def person_options(self, include_everyone: bool = False) -> list[str]:
        configured = list(self.config_data.get("select_options", {}).get("İsim", []))
        existing = sorted({row.get("İsim", "") for row in self.rows if row.get("İsim", "")})
        options = sorted(set(configured + existing))
        if include_everyone:
            return ["Herkes"] + [item for item in options if normalize_search(item) != normalize_search("Herkes")]
        return options

    def refresh_kpi_person_options(self) -> None:
        if hasattr(self, "kpi_person_combo"):
            self.kpi_person_combo.configure(values=self.person_options(include_everyone=True))
        if hasattr(self, "filter_person_combo"):
            self.filter_person_combo.configure(values=self.person_options(include_everyone=True))
        if (
            hasattr(self, "kpi_person_count_var")
            and (not self.kpi_person_count_var.get().strip() or self.is_everyone_selection(self.kpi_person_var.get()))
        ):
            self.kpi_person_count_var.set(str(self.default_kpi_person_count()))

    def attach_person_autocomplete(self, combo: ttk.Combobox, variable: tk.StringVar, include_everyone: bool = False) -> None:
        def matches(text: str) -> list[str]:
            needle = normalize_search(text.strip())
            options = self.person_options(include_everyone=include_everyone)
            if not needle:
                return options
            return [item for item in options if needle in normalize_search(item)]

        def on_keyrelease(event) -> None:
            if event.keysym in {"Up", "Down", "Left", "Right", "Tab", "Escape"}:
                return
            current = variable.get()
            suggestions = matches(current)
            combo.configure(values=suggestions)
            if suggestions:
                combo.event_generate("<Down>")

        def on_return(_event) -> str:
            suggestions = matches(variable.get())
            if suggestions:
                variable.set(suggestions[0])
                combo.icursor(tk.END)
            return "break"

        combo.bind("<KeyRelease>", on_keyrelease)
        combo.bind("<Return>", on_return)

    def create_summary(self, parent: ttk.Frame) -> None:
        summary = ttk.LabelFrame(parent, text=self.label("totals", "Toplamlar"))
        summary.pack(fill="x", padx=4, pady=(10, 4))
        summary_fields = self.config_data.get("summary_fields", {})
        items = [
            (summary_fields.get("izin", "Toplam İzin"), "izin"),
            (summary_fields.get("mesai", "Toplam Mesai"), "mesai"),
            (summary_fields.get("rapor", "Toplam Rapor"), "rapor"),
            (summary_fields.get("resmi", "Resmi Tatil Mesaisi"), "resmi"),
        ]
        for index, (label, key) in enumerate(items):
            ttk.Label(summary, text=label).grid(row=0, column=index * 2, sticky="e", padx=(12, 4), pady=10)
            ttk.Label(summary, textvariable=self.summary_vars[key], font=("", 11, "bold")).grid(
                row=0,
                column=index * 2 + 1,
                sticky="w",
                padx=(0, 18),
                pady=10,
            )

    def create_scrollable_settings_area(self, parent: ttk.Frame) -> ttk.Frame:
        canvas = tk.Canvas(parent, bg=self.color("background_color"), highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        content = ttk.Frame(canvas)
        content.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        return content

    def create_personnel_settings_tab(self, parent: ttk.Frame) -> None:
        content = self.create_scrollable_settings_area(parent)
        self.create_record_settings(content)
        self.create_holiday_settings(content)

    def create_general_settings_tab(self, parent: ttk.Frame) -> None:
        content = self.create_scrollable_settings_area(parent)
        self.create_general_settings(content)
        self.create_work_settings(content)

    def setting_entry(self, parent: ttk.Frame, label: str, value: str, row: int, column: int = 0, width: int = 18) -> tk.StringVar:
        variable = tk.StringVar(value=value)
        ttk.Label(parent, text=label).grid(row=row, column=column, sticky="w", padx=6, pady=(6, 2))
        ttk.Entry(parent, textvariable=variable, width=width).grid(row=row + 1, column=column, sticky="ew", padx=6, pady=(0, 6))
        parent.columnconfigure(column, weight=1)
        return variable

    def setting_check(self, parent: ttk.Frame, label: str, value: bool, row: int, column: int = 0) -> tk.BooleanVar:
        variable = tk.BooleanVar(value=value)
        ttk.Checkbutton(parent, text=label, variable=variable).grid(row=row, column=column, sticky="w", padx=6, pady=6)
        return variable

    def setting_color_entry(self, parent: ttk.Frame, label: str, value: str, row: int, column: int = 0) -> tk.StringVar:
        variable = tk.StringVar(value=value)
        ttk.Label(parent, text=label).grid(row=row, column=column, sticky="w", padx=6, pady=(6, 2))
        field_frame = ttk.Frame(parent)
        field_frame.grid(row=row + 1, column=column, sticky="ew", padx=6, pady=(0, 6))
        entry = ttk.Entry(field_frame, textvariable=variable, width=13)
        entry.pack(side="left", fill="x", expand=True)
        preview = tk.Label(field_frame, text="", width=3, relief="solid", bg=value)
        preview.pack(side="left", padx=(6, 0))

        def update_preview(*_args) -> None:
            color = variable.get().strip()
            try:
                preview.configure(bg=color)
            except tk.TclError:
                preview.configure(bg=self.color("background_color"))

        def choose_color(_event=None) -> None:
            _rgb, hex_color = colorchooser.askcolor(color=variable.get().strip() or value, parent=self)
            if hex_color:
                variable.set(hex_color.upper())
                update_preview()

        preview.bind("<Button-1>", choose_color)
        entry.bind("<Button-1>", choose_color)
        variable.trace_add("write", update_preview)
        parent.columnconfigure(column, weight=1)
        return variable

    def ask_settings_password(self) -> bool:
        password = simpledialog.askstring("Ayar Şifresi", "Ayarları kaydetmek için şifre girin:", show="*", parent=self)
        if password is None:
            return False
        if password != SETTINGS_PASSWORD:
            messagebox.showwarning("Hatalı Şifre", "Şifre yanlış. Ayarlar kaydedilmedi.")
            return False
        return True

    def save_runtime_config(self, require_password: bool = True) -> bool:
        if require_password and not self.ask_settings_password():
            return False
        save_config(self.config_data)
        self.theme = self.config_data.get("theme", {})
        self.date_settings = self.config_data.get("date_settings", {})
        self.date_format = self.date_settings.get("display_format", "dd.MM.yyyy")
        self.date_strftime = date_format_to_strftime(self.date_format)
        self.rules = rules_from_config(self.config_data)
        self.title(self.config_data["program"]["name"])
        self.apply_theme()
        self.update_logo_display()
        for column, widget in self.combo_widgets.items():
            widget.configure(values=self.config_data.get("select_options", {}).get(column, []))
        self.refresh_kpi_person_options()
        self.refresh_table()
        return True

    def create_general_settings(self, parent: ttk.Frame) -> None:
        section = ttk.LabelFrame(parent, text="Genel Ayarlar")
        section.pack(fill="x", padx=8, pady=8)

        program = self.config_data["program"]
        logo = self.config_data.setdefault("logo", {})
        kpi_settings = self.config_data.setdefault("kpi_settings", {})
        theme = self.config_data["theme"]
        date_settings = self.config_data["date_settings"]

        name_var = self.setting_entry(section, "Program adı", program.get("name", ""), 0, 0)
        kpi_target_var = self.setting_entry(section, "KPI aylık hedef saat", str(kpi_settings.get("monthly_required_work_hours", 270)), 0, 1)
        date_format_var = self.setting_entry(section, "Tarih formatı", date_settings.get("display_format", "dd.MM.yyyy"), 0, 2)
        default_today_var = self.setting_check(section, "Açılışta bugünün tarihi gelsin", bool(date_settings.get("default_today_on_start", True)), 0, 3)
        calendar_var = self.setting_check(section, "Tarih alanında takvim açılsın", bool(date_settings.get("use_calendar_picker", True)), 1, 3)

        logo_text_var = self.setting_entry(section, "Logo yazısı", logo.get("text", "HS"), 2, 0)
        logo_font_var = self.setting_entry(section, "Logo fontu", logo.get("font", "Segoe UI Semibold"), 2, 1)
        logo_h_color_var = self.setting_color_entry(section, "Logo H rengi", logo.get("h_color", "#001F5B"), 2, 2)
        logo_s_color_var = self.setting_color_entry(section, "Logo S rengi", logo.get("s_color", "#000000"), 2, 3)

        theme_keys = [
            ("primary_color", "Ana renk"),
            ("secondary_color", "İkinci renk"),
            ("background_color", "Arka plan"),
            ("text_color", "Yazı rengi"),
            ("button_color", "Buton rengi"),
            ("button_text_color", "Buton yazı rengi"),
            ("input_background_color", "Giriş alanı"),
            ("table_background_color", "Tablo arka plan"),
            ("card_background_color", "Kart arka plan"),
            ("chart_background_color", "Grafik arka plan"),
            ("accent_color", "Vurgu rengi"),
            ("zebra_color", "Zebra satır"),
        ]
        theme_vars = {}
        for index, (key, label) in enumerate(theme_keys):
            row = 4 + (index // 4) * 2
            column = index % 4
            theme_vars[key] = self.setting_color_entry(section, label, theme.get(key, ""), row, column)
        after_theme_row = 4 + ((len(theme_keys) + 3) // 4) * 2

        def save_general() -> None:
            if not self.ask_settings_password():
                return
            try:
                kpi_settings["monthly_required_work_hours"] = float(kpi_target_var.get().strip().replace(",", "."))
            except ValueError:
                messagebox.showwarning("KPI", "KPI aylık hedef saat sayısal olmalı.")
                return
            program["name"] = name_var.get().strip()
            logo["text"] = logo_text_var.get().strip() or "HS"
            logo["font"] = logo_font_var.get().strip() or "Segoe UI Semibold"
            logo["h_color"] = logo_h_color_var.get().strip() or "#001F5B"
            logo["s_color"] = logo_s_color_var.get().strip() or "#000000"
            date_settings["display_format"] = date_format_var.get().strip()
            date_settings["default_today_on_start"] = bool(default_today_var.get())
            date_settings["use_calendar_picker"] = bool(calendar_var.get())
            for key, variable in theme_vars.items():
                theme[key] = variable.get().strip()
            self.save_runtime_config(require_password=False)
            messagebox.showinfo("Ayarlar", "Genel ayarlar kaydedildi.")

        ttk.Button(section, text="Kaydet", command=save_general).grid(row=after_theme_row, column=0, sticky="w", padx=6, pady=8)
        self.create_list_editor(section, "Arama seçenekleri", self.config_data["search_fields"], after_theme_row + 1, 0, 4, self.save_runtime_config)

    def create_record_settings(self, parent: ttk.Frame) -> None:
        section = ttk.LabelFrame(parent, text="Personel ve Kayıt Ayarları")
        section.pack(fill="x", padx=8, pady=8)
        options = self.config_data.setdefault("select_options", {})
        editors = [
            ("Personel isimleri", options.setdefault("İsim", [])),
            ("Departmanlar", options.setdefault("Departman", [])),
            ("Kayıt türleri", options.setdefault("Tür", [])),
            ("Açıklama seçenekleri", options.setdefault("Açıklama", [])),
        ]
        for index, (title, values) in enumerate(editors):
            self.create_list_editor(section, title, values, 0, index, 1, lambda: self.save_runtime_config(require_password=False))

    def create_work_settings(self, parent: ttk.Frame) -> None:
        section = ttk.LabelFrame(parent, text="Çalışma Saatleri")
        section.pack(fill="x", padx=8, pady=8)
        work_hours = self.config_data["work_hours"]

        weekday_start = self.setting_entry(section, "Hafta içi başlangıç", work_hours.get("weekday_start", ""), 0, 0)
        weekday_end = self.setting_entry(section, "Hafta içi bitiş", work_hours.get("weekday_end", ""), 0, 1)
        saturday_start = self.setting_entry(section, "Cumartesi başlangıç", work_hours.get("saturday_start", ""), 0, 2)
        saturday_end = self.setting_entry(section, "Cumartesi bitiş", work_hours.get("saturday_end", ""), 0, 3)
        saturday_enabled = self.setting_check(section, "Cumartesi çalışılır", bool(work_hours.get("saturday_enabled", True)), 2, 0)
        sunday_enabled = self.setting_check(section, "Pazar çalışılır", bool(work_hours.get("sunday_enabled", False)), 2, 1)
        sunday_start = self.setting_entry(section, "Pazar başlangıç", work_hours.get("sunday_start", ""), 3, 0)
        sunday_end = self.setting_entry(section, "Pazar bitiş", work_hours.get("sunday_end", ""), 3, 1)

        def save_work_hours() -> None:
            if not self.ask_settings_password():
                return
            work_hours["weekday_start"] = weekday_start.get().strip()
            work_hours["weekday_end"] = weekday_end.get().strip()
            work_hours["saturday_start"] = saturday_start.get().strip()
            work_hours["saturday_end"] = saturday_end.get().strip()
            work_hours["saturday_enabled"] = bool(saturday_enabled.get())
            work_hours["sunday_enabled"] = bool(sunday_enabled.get())
            work_hours["sunday_start"] = sunday_start.get().strip()
            work_hours["sunday_end"] = sunday_end.get().strip()
            self.save_runtime_config(require_password=False)
            messagebox.showinfo("Ayarlar", "Çalışma saatleri kaydedildi.")

        ttk.Button(section, text="Kaydet", command=save_work_hours).grid(row=5, column=0, sticky="w", padx=6, pady=8)
        self.create_break_editor(section, 6)

    def create_holiday_settings(self, parent: ttk.Frame) -> None:
        section = ttk.LabelFrame(parent, text="Tatil Ayarları")
        section.pack(fill="x", padx=8, pady=8)
        self.create_list_editor(section, "Resmi tatiller", self.config_data["holidays"]["full_days"], 0, 0, 1, lambda: self.save_runtime_config(require_password=False))
        self.create_half_holiday_editor(section, 0, 1)

    def create_list_editor(self, parent: ttk.Frame, title: str, values: list[str], row: int, column: int, columnspan: int, on_save) -> None:
        frame = ttk.LabelFrame(parent, text=title)
        frame.grid(row=row, column=column, columnspan=columnspan, sticky="nsew", padx=6, pady=6)
        parent.columnconfigure(column, weight=1)
        saved_values = list(values)
        listbox = tk.Listbox(frame, height=6, bg=self.color("table_background_color"), fg=self.color("text_color"))
        listbox.grid(row=0, column=0, columnspan=4, sticky="nsew", padx=6, pady=6)
        value_var = tk.StringVar()
        ttk.Entry(frame, textvariable=value_var).grid(row=1, column=0, columnspan=4, sticky="ew", padx=6, pady=(0, 6))

        def refresh() -> None:
            listbox.delete(0, tk.END)
            for item in values:
                listbox.insert(tk.END, item)

        def load_selected(_event=None) -> None:
            selection = listbox.curselection()
            if selection:
                value_var.set(values[selection[0]])

        def add() -> None:
            value = value_var.get().strip()
            if value and value not in values:
                values.append(value)
                refresh()

        def update() -> None:
            selection = listbox.curselection()
            value = value_var.get().strip()
            if selection and value:
                values[selection[0]] = value
                refresh()

        def delete() -> None:
            selection = listbox.curselection()
            if selection:
                del values[selection[0]]
                value_var.set("")
                refresh()

        def save() -> None:
            if on_save():
                saved_values[:] = list(values)
                messagebox.showinfo("Ayarlar", f"{title} kaydedildi.")
            else:
                values[:] = saved_values
                refresh()

        listbox.bind("<<ListboxSelect>>", load_selected)
        ttk.Button(frame, text="Ekle", command=add).grid(row=2, column=0, padx=3, pady=6)
        ttk.Button(frame, text="Düzenle", command=update).grid(row=2, column=1, padx=3, pady=6)
        ttk.Button(frame, text="Sil", command=delete).grid(row=2, column=2, padx=3, pady=6)
        ttk.Button(frame, text="Kaydet", command=save).grid(row=2, column=3, padx=3, pady=6)
        frame.columnconfigure(0, weight=1)
        refresh()

    def create_break_editor(self, parent: ttk.Frame, row: int) -> None:
        frame = ttk.LabelFrame(parent, text="Mola Saatleri")
        frame.grid(row=row, column=0, columnspan=4, sticky="ew", padx=6, pady=6)
        breaks = self.config_data["breaks"]
        saved_breaks = [dict(item) for item in breaks]
        listbox = tk.Listbox(frame, height=5, bg=self.color("table_background_color"), fg=self.color("text_color"))
        listbox.grid(row=0, column=0, columnspan=5, sticky="ew", padx=6, pady=6)
        name_var = self.setting_entry(frame, "Ad", "", 1, 0)
        start_var = self.setting_entry(frame, "Başlangıç", "", 1, 1)
        end_var = self.setting_entry(frame, "Bitiş", "", 1, 2)
        applies_var = self.setting_entry(frame, "Günler", "weekday", 1, 3)

        def refresh() -> None:
            listbox.delete(0, tk.END)
            for item in breaks:
                listbox.insert(tk.END, f"{item.get('name', '')} | {item.get('start', '')}-{item.get('end', '')} | {','.join(item.get('applies_to', []))}")

        def load_selected(_event=None) -> None:
            selection = listbox.curselection()
            if not selection:
                return
            item = breaks[selection[0]]
            name_var.set(item.get("name", ""))
            start_var.set(item.get("start", ""))
            end_var.set(item.get("end", ""))
            applies_var.set(",".join(item.get("applies_to", [])))

        def current_item() -> dict:
            return {
                "name": name_var.get().strip(),
                "start": start_var.get().strip(),
                "end": end_var.get().strip(),
                "applies_to": [item.strip() for item in applies_var.get().split(",") if item.strip()],
            }

        def add() -> None:
            breaks.append(current_item())
            refresh()

        def update() -> None:
            selection = listbox.curselection()
            if selection:
                breaks[selection[0]] = current_item()
                refresh()

        def delete() -> None:
            selection = listbox.curselection()
            if selection:
                del breaks[selection[0]]
                refresh()

        def save() -> None:
            if self.save_runtime_config():
                saved_breaks[:] = [dict(item) for item in breaks]
                messagebox.showinfo("Ayarlar", "Mola saatleri kaydedildi.")
            else:
                breaks[:] = [dict(item) for item in saved_breaks]
                refresh()

        listbox.bind("<<ListboxSelect>>", load_selected)
        ttk.Button(frame, text="Ekle", command=add).grid(row=3, column=0, padx=3, pady=6)
        ttk.Button(frame, text="Düzenle", command=update).grid(row=3, column=1, padx=3, pady=6)
        ttk.Button(frame, text="Sil", command=delete).grid(row=3, column=2, padx=3, pady=6)
        ttk.Button(frame, text="Kaydet", command=save).grid(row=3, column=3, padx=3, pady=6)
        refresh()

    def create_half_holiday_editor(self, parent: ttk.Frame, row: int, column: int) -> None:
        frame = ttk.LabelFrame(parent, text="Yarım gün tatiller")
        frame.grid(row=row, column=column, sticky="nsew", padx=6, pady=6)
        half_days = self.config_data["holidays"]["half_days"]
        listbox = tk.Listbox(frame, height=6, bg=self.color("table_background_color"), fg=self.color("text_color"))
        listbox.grid(row=0, column=0, columnspan=4, sticky="ew", padx=6, pady=6)
        date_var = self.setting_entry(frame, "Tarih", "", 1, 0)
        start_var = self.setting_entry(frame, "Başlangıç", "", 1, 1)
        end_var = self.setting_entry(frame, "Bitiş", "", 1, 2)

        def refresh() -> None:
            listbox.delete(0, tk.END)
            for item in half_days:
                listbox.insert(tk.END, f"{item.get('date', '')} | {item.get('start', '')}-{item.get('end', '')}")

        def load_selected(_event=None) -> None:
            selection = listbox.curselection()
            if not selection:
                return
            item = half_days[selection[0]]
            date_var.set(item.get("date", ""))
            start_var.set(item.get("start", ""))
            end_var.set(item.get("end", ""))

        def current_item() -> dict:
            return {"date": date_var.get().strip(), "start": start_var.get().strip(), "end": end_var.get().strip()}

        def add() -> None:
            half_days.append(current_item())
            refresh()

        def update() -> None:
            selection = listbox.curselection()
            if selection:
                half_days[selection[0]] = current_item()
                refresh()

        def delete() -> None:
            selection = listbox.curselection()
            if selection:
                del half_days[selection[0]]
                refresh()

        def save() -> None:
            if self.save_runtime_config(require_password=False):
                messagebox.showinfo("Ayarlar", "Yarım gün tatiller kaydedildi.")

        listbox.bind("<<ListboxSelect>>", load_selected)
        ttk.Button(frame, text="Ekle", command=add).grid(row=3, column=0, padx=3, pady=6)
        ttk.Button(frame, text="Düzenle", command=update).grid(row=3, column=1, padx=3, pady=6)
        ttk.Button(frame, text="Sil", command=delete).grid(row=3, column=2, padx=3, pady=6)
        ttk.Button(frame, text="Kaydet", command=save).grid(row=3, column=3, padx=3, pady=6)
        refresh()

    def read_form_row(self) -> dict[str, str]:
        row = {column: self.entry_vars[column].get().strip() for column in self.input_columns}
        if any(not row.get(column, "") for column in self.input_columns if column != "Açıklama"):
            raise ValueError("Zorunlu kayıt alanları doldurulmalı.")
        calculate_row(row, self.rules)
        return row

    def set_default_dates(self) -> None:
        if not self.date_settings.get("default_today_on_start", True):
            return
        today_text = self.format_date(date.today())
        for column in self.input_columns:
            if self.is_date_column(column) and not self.entry_vars[column].get().strip():
                self.entry_vars[column].set(today_text)

    def open_calendar(self, column: str) -> None:
        current_value = self.entry_vars[column].get().strip()
        try:
            selected = parse_date(current_value, self.date_format) if current_value else date.today()
        except ValueError:
            selected = date.today()
        CalendarPopup(self, selected, lambda value: self.entry_vars[column].set(self.format_date(value)), self.theme)

    def fill_form(self, row: dict[str, str]) -> None:
        for column in self.input_columns:
            self.entry_vars[column].set(row.get(column, ""))

    def add_row(self) -> None:
        try:
            row = self.read_form_row()
        except Exception as exc:
            messagebox.showerror("Kayıt eklenemedi", str(exc))
            return
        self.rows.append(row)
        self.refresh_kpi_person_options()
        self.refresh_table()

    def delete_selected(self) -> None:
        selected = self.tree.selection()
        indexes = sorted((int(item) for item in selected), reverse=True)
        for index in indexes:
            if 0 <= index < len(self.rows):
                del self.rows[index]
        self.refresh_kpi_person_options()
        self.refresh_table()

    def display_row(self, row: dict[str, str]) -> dict[str, str]:
        calculated = calculate_row(row, self.rules)
        display = {column: row.get(column, "") for column in self.columns}
        for column, key in CALCULATED_KEYS.items():
            if column in display:
                display[column] = format_hours(calculated[key])
        return display

    def refresh_kpi(self) -> None:
        if self.kpi_running:
            return
        person = self.kpi_person_var.get().strip() or "Herkes"
        if not person:
            messagebox.showwarning("KPI", "Lütfen personel seçin.")
            return
        try:
            month = int(self.kpi_month_var.get())
            year = int(self.kpi_year_var.get())
            person_count = int(float(self.kpi_person_count_var.get().strip().replace(",", ".")))
        except ValueError:
            messagebox.showwarning("KPI", "Ay, yıl ve personel sayısı geçerli olmalı.")
            return
        if person_count < 0:
            messagebox.showwarning("KPI", "Personel sayısı negatif olamaz.")
            return
        self.kpi_running = True
        self.kpi_status_var.set("KPI hesaplanıyor...")
        thread = threading.Thread(target=self.kpi_worker, args=(person, month, year, person_count), daemon=True)
        thread.start()
        self.after(100, self.poll_kpi_queue)

    def month_bounds(self, month: int, year: int) -> tuple[datetime, datetime]:
        start = datetime(year, month, 1)
        if month == 12:
            end = datetime(year + 1, 1, 1)
        else:
            end = datetime(year, month + 1, 1)
        return start, end

    def month_capacity_minutes_per_person(self, month: int, year: int) -> tuple[int, int]:
        first_day = date(year, month, 1)
        last_day_number = calendar.monthrange(year, month)[1]
        total_minutes = 0
        workday_count = 0
        for day_number in range(1, last_day_number + 1):
            current_day = first_day.replace(day=day_number)
            day_minutes = sum(minutes_between(start, end) for start, end in working_periods(current_day, self.rules))
            if day_minutes > 0:
                workday_count += 1
                total_minutes += day_minutes
        return total_minutes, workday_count

    def format_hours_with_days(self, minutes: int, day_minutes: float) -> str:
        if day_minutes <= 0:
            return f"{format_hours(minutes)} / 0 gün"
        days_value = minutes / day_minutes
        if abs(days_value - round(days_value)) < 0.01:
            day_text = str(int(round(days_value)))
        else:
            day_text = f"{days_value:.1f}"
        return f"{format_hours(minutes)} / {day_text} gün"

    def clipped_row_for_month(self, row: dict[str, str], month: int, year: int) -> dict[str, str] | None:
        start = datetime.combine(parse_date(row["Başlangıç Tarihi"], self.date_format), datetime.strptime(row["Başlangıç Saati"], "%H:%M").time())
        end = datetime.combine(parse_date(row["Bitiş Tarihi"], self.date_format), datetime.strptime(row["Bitiş Saati"], "%H:%M").time())
        month_start, month_end = self.month_bounds(month, year)
        clipped_start = max(start, month_start)
        clipped_end = min(end, month_end)
        if clipped_start >= clipped_end:
            return None
        clipped = dict(row)
        clipped["Başlangıç Tarihi"] = self.format_date(clipped_start.date())
        clipped["Başlangıç Saati"] = clipped_start.strftime("%H:%M")
        clipped["Bitiş Tarihi"] = self.format_date(clipped_end.date())
        clipped["Bitiş Saati"] = clipped_end.strftime("%H:%M")
        return clipped

    def is_everyone_selection(self, person: str) -> bool:
        return normalize_search(person) == normalize_search("Herkes")

    def row_matches_kpi_person(self, row: dict[str, str], person: str) -> bool:
        return self.is_everyone_selection(person) or row.get("İsim", "") == person

    def row_loss_minutes(self, row: dict[str, str], calculated: dict[str, int]) -> tuple[int, str | None]:
        kind = normalize_kind(row.get("Tür", ""))
        if kind == "izin":
            return calculated["izin_dk"], "izin"
        if kind in {"rapor", "raporlu"}:
            return calculated["rapor_dk"], "rapor"
        if kind in {"mesai", "fazla mesai", "resmi tatil mesai", "resmi tatilde mesai"}:
            return 0, None

        start = datetime.combine(parse_date(row["Başlangıç Tarihi"], self.date_format), datetime.strptime(row["Başlangıç Saati"], "%H:%M").time())
        end = datetime.combine(parse_date(row["Bitiş Tarihi"], self.date_format), datetime.strptime(row["Bitiş Saati"], "%H:%M").time())
        loss_minutes = 0
        current_day = start.date()
        while current_day <= end.date():
            day_start = datetime.combine(current_day, time.min)
            day_end = day_start + timedelta(days=1)
            current_start = max(start, day_start)
            current_end = min(end, day_end)
            if current_start < current_end:
                loss_minutes += sum_overlaps(current_start, current_end, working_periods(current_day, self.rules))
            current_day += timedelta(days=1)
        return loss_minutes, "diger"

    def kpi_worker(self, person: str, month: int, year: int, person_count: int) -> None:
        try:
            totals = {
                "fazla_mesai": 0,
                "izin": 0,
                "rapor": 0,
                "resmi": 0,
                "toplam": 0,
                "toplam_kayip": 0,
                "toplam_kapasite": 0,
                "izin_kayip": 0,
                "rapor_kayip": 0,
                "diger_kayip": 0,
                "planlanan": 0,
                "kullanilabilir": 0,
                "ek_kapasite": 0,
                "eksik_oran": 0.0,
                "mesai_limit": 0,
                "kalan_limit": 0,
                "limit_oran": 0.0,
                "day_minutes": 0,
            }
            by_day: dict[int, dict[str, int]] = {}
            monthly_line = [0 for _ in range(12)]
            top_by_person: dict[str, dict[str, int]] = {}
            mesai_limit = int(float(self.config_data.get("kpi_settings", {}).get("monthly_required_work_hours", 270)) * 60)
            selected_limit_used = 0

            for row in list(self.rows):
                clipped = self.clipped_row_for_month(row, month, year)
                if not clipped:
                    continue
                try:
                    calculated = calculate_row(clipped, self.rules)
                except ValueError:
                    calculated = {
                        "mesai_dk": 0,
                        "izin_dk": 0,
                        "rapor_dk": 0,
                        "resmi_tatil_mesai_dk": 0,
                        "toplam_dk": 0,
                    }

                person_name = row.get("İsim", "")
                person_bucket = top_by_person.setdefault(person_name, {"fazla_mesai": 0, "resmi": 0})
                person_bucket["fazla_mesai"] += calculated["mesai_dk"]
                person_bucket["resmi"] += calculated["resmi_tatil_mesai_dk"]

                if not self.row_matches_kpi_person(row, person):
                    continue

                totals["fazla_mesai"] += calculated["mesai_dk"]
                totals["izin"] += calculated["izin_dk"]
                totals["rapor"] += calculated["rapor_dk"]
                totals["resmi"] += calculated["resmi_tatil_mesai_dk"]
                totals["toplam"] += calculated["toplam_dk"]
                loss_minutes, loss_kind = self.row_loss_minutes(clipped, calculated)
                totals["toplam_kayip"] += loss_minutes
                if loss_kind == "izin":
                    totals["izin_kayip"] += loss_minutes
                elif loss_kind == "rapor":
                    totals["rapor_kayip"] += loss_minutes
                elif loss_kind == "diger":
                    totals["diger_kayip"] += loss_minutes

                start_day = parse_date(clipped["Başlangıç Tarihi"], self.date_format).day
                day_bucket = by_day.setdefault(start_day, {"mesai": 0, "izin": 0, "rapor": 0, "resmi": 0, "kayip": 0})
                day_bucket["mesai"] += calculated["mesai_dk"]
                day_bucket["izin"] += calculated["izin_dk"]
                day_bucket["rapor"] += calculated["rapor_dk"]
                day_bucket["resmi"] += calculated["resmi_tatil_mesai_dk"]
                day_bucket["kayip"] += loss_minutes
                selected_limit_used += calculated["mesai_dk"] + calculated["resmi_tatil_mesai_dk"]

            for line_month in range(1, 13):
                for row in list(self.rows):
                    if not self.row_matches_kpi_person(row, person):
                        continue
                    clipped = self.clipped_row_for_month(row, line_month, year)
                    if clipped:
                        try:
                            calculated = calculate_row(clipped, self.rules)
                        except ValueError:
                            continue
                        monthly_line[line_month - 1] += calculated["mesai_dk"] + calculated["resmi_tatil_mesai_dk"]

            capacity_person_count = person_count if self.is_everyone_selection(person) else 1
            capacity_per_person, workday_count = self.month_capacity_minutes_per_person(month, year)
            totals["day_minutes"] = int(capacity_per_person / workday_count) if workday_count else 0
            totals["planlanan"] = capacity_per_person * capacity_person_count
            totals["toplam_kapasite"] = totals["planlanan"]
            totals["toplam_kayip"] = totals["izin"] + totals["rapor"]
            totals["izin_kayip"] = totals["izin"]
            totals["rapor_kayip"] = totals["rapor"]
            totals["kullanilabilir"] = totals["planlanan"] - totals["izin"] - totals["rapor"]
            totals["ek_kapasite"] = totals["fazla_mesai"] + totals["resmi"]
            totals["eksik_oran"] = ((totals["izin"] + totals["rapor"]) / totals["planlanan"] * 100) if totals["planlanan"] > 0 else 0.0
            if self.is_everyone_selection(person):
                total_overtime_all = sum(values["fazla_mesai"] + values["resmi"] for values in top_by_person.values())
                selected_limit_used = int(total_overtime_all / person_count) if person_count > 0 else 0
            totals["mesai_limit"] = mesai_limit
            totals["kalan_limit"] = max(0, mesai_limit - selected_limit_used)
            totals["limit_oran"] = (selected_limit_used / mesai_limit * 100) if mesai_limit > 0 else 0.0

            top_rows = []
            for person_name, values in top_by_person.items():
                if not person_name:
                    continue
                total_overtime = values["fazla_mesai"] + values["resmi"]
                top_rows.append(
                    {
                        "person": person_name,
                        "fazla_mesai": values["fazla_mesai"],
                        "resmi": values["resmi"],
                        "toplam": total_overtime,
                        "limit_oran": (total_overtime / mesai_limit * 100) if mesai_limit > 0 else 0.0,
                    }
                )
            top_rows.sort(key=lambda item: item["toplam"], reverse=True)
            self.kpi_queue.put(("done", {"totals": totals, "by_day": by_day, "monthly_line": monthly_line, "top_rows": top_rows}))
        except Exception as exc:
            self.kpi_queue.put(("error", str(exc)))

    def poll_kpi_queue(self) -> None:
        try:
            message_type, payload = self.kpi_queue.get_nowait()
        except queue.Empty:
            if self.kpi_running:
                self.after(100, self.poll_kpi_queue)
            return
        self.kpi_running = False
        if message_type == "error":
            self.kpi_status_var.set("KPI hesaplanamadı")
            messagebox.showerror("KPI", payload)
            return
        self.kpi_status_var.set("KPI hazır")
        self.render_kpi(payload)

    def render_kpi(self, data: dict) -> None:
        totals = data["totals"]
        day_minutes = totals.get("day_minutes", 0)
        self.kpi_card_vars["fazla_mesai"].set(self.format_hours_with_days(totals["fazla_mesai"], day_minutes))
        self.kpi_card_vars["izin"].set(self.format_hours_with_days(totals["izin"], day_minutes))
        self.kpi_card_vars["rapor"].set(self.format_hours_with_days(totals["rapor"], day_minutes))
        self.kpi_card_vars["resmi"].set(self.format_hours_with_days(totals["resmi"], day_minutes))
        self.kpi_card_vars["toplam_kayip"].set(self.format_hours_with_days(totals["toplam_kayip"], day_minutes))
        self.kpi_card_vars["toplam_kapasite"].set(self.format_hours_with_days(totals["toplam_kapasite"], day_minutes))
        self.kpi_card_vars["kullanilabilir"].set(self.format_hours_with_days(totals["kullanilabilir"], day_minutes))
        self.kpi_card_vars["ek_kapasite"].set(self.format_hours_with_days(totals["ek_kapasite"], day_minutes))
        self.kpi_card_vars["eksik_oran"].set(f"%{totals['eksik_oran']:.2f}")
        self.kpi_card_vars["kalan_limit"].set(format_hours(totals["kalan_limit"]))
        if totals["limit_oran"] >= 90:
            self.kpi_status_var.set("Mesai limiti uyarısı: %90 üzeri")
        elif totals["limit_oran"] >= 75:
            self.kpi_status_var.set("Mesai limiti yaklaşıyor")
        self.draw_loss_chart(totals)
        self.draw_extra_capacity_chart(totals)
        self.render_top_overtime(data.get("top_rows", []))

    def render_top_overtime(self, rows: list[dict]) -> None:
        if not hasattr(self, "top_overtime_tree"):
            return
        for item in self.top_overtime_tree.get_children():
            self.top_overtime_tree.delete(item)
        for index, row in enumerate(rows):
            tag = "even" if index % 2 else "odd"
            self.top_overtime_tree.insert(
                "",
                "end",
                values=(
                    row["person"],
                    format_hours(row["fazla_mesai"]),
                    format_hours(row["resmi"]),
                    format_hours(row["toplam"]),
                    f"%{row['limit_oran']:.2f}",
                ),
                tags=(tag,),
            )

    def current_kpi_report_payload(self) -> tuple[str, int, int, int, dict]:
        person = self.kpi_person_var.get().strip() or "Herkes"
        month = int(self.kpi_month_var.get())
        year = int(self.kpi_year_var.get())
        person_count = int(float(self.kpi_person_count_var.get().strip().replace(",", ".")))
        old_queue = self.kpi_queue
        self.kpi_queue = queue.Queue()
        try:
            self.kpi_worker(person, month, year, person_count)
            message_type, payload = self.kpi_queue.get_nowait()
        finally:
            self.kpi_queue = old_queue
        if message_type == "error":
            raise ValueError(payload)
        return person, month, year, person_count, payload

    def kpi_report_cards(self, totals: dict) -> list[tuple[str, str]]:
        day_minutes = totals.get("day_minutes", 0)
        return [
            ("Fazla Mesai Saati", self.format_hours_with_days(totals["fazla_mesai"], day_minutes)),
            ("Kullanılan İzin Saati", self.format_hours_with_days(totals["izin"], day_minutes)),
            ("Rapor Saati", self.format_hours_with_days(totals["rapor"], day_minutes)),
            ("Resmi Tatil Mesai Saati", self.format_hours_with_days(totals["resmi"], day_minutes)),
            ("Toplam İşçilik Kaybı", self.format_hours_with_days(totals["toplam_kayip"], day_minutes)),
            ("Toplam Kapasite", self.format_hours_with_days(totals["toplam_kapasite"], day_minutes)),
            ("Eksik Çalışmalar Hariç Kapasite", self.format_hours_with_days(totals["kullanilabilir"], day_minutes)),
            ("Fazla Mesai ile Gelen Ek Kapasite", self.format_hours_with_days(totals["ek_kapasite"], day_minutes)),
            ("Eksik Çalışma Oranı %", f"%{totals['eksik_oran']:.2f}"),
            ("Kişi Bazlı Kalan Mesai Limiti", format_hours(totals["kalan_limit"])),
        ]

    def kpi_raw_rows_for_report(self, person: str, month: int, year: int) -> list[dict[str, str]]:
        rows = []
        for row in self.rows:
            if not self.row_matches_kpi_person(row, person):
                continue
            if self.clipped_row_for_month(row, month, year):
                rows.append(self.personnel_export_row(row))
        return rows

    def draw_report_pie_png(self, path: Path, title: str, values: list[tuple[str, int, str]]) -> None:
        from PIL import Image, ImageDraw, ImageFont

        image = Image.new("RGB", (760, 420), self.color("chart_background_color", "#FFFFFF"))
        draw = ImageDraw.Draw(image)
        try:
            font_title = ImageFont.truetype("arial.ttf", 24)
            font = ImageFont.truetype("arial.ttf", 17)
            font_bold = ImageFont.truetype("arialbd.ttf", 18)
        except OSError:
            font_title = font = font_bold = ImageFont.load_default()
        draw.text((28, 24), title, fill=self.color("text_color"), font=font_title)
        total = sum(value for _label, value, _color in values)
        if total <= 0:
            draw.text((330, 200), "Veri yok", fill=self.color("text_color"), font=font)
            image.save(path)
            return
        box = (70, 85, 350, 365)
        start = 0.0
        for label, value, color in values:
            extent = value / total * 360
            draw.pieslice(box, start=start, end=start + extent, fill=color, outline="white")
            if value > 0:
                angle = math.radians(start + extent / 2)
                x = 210 + 92 * math.cos(angle)
                y = 225 + 92 * math.sin(angle)
                percent_text = f"%{value / total * 100:.1f}"
                draw.text((x - 24, y - 10), percent_text, fill="white", font=font_bold)
            start += extent
        legend_x = 410
        for index, (label, value, color) in enumerate(values):
            y = 120 + index * 44
            draw.rectangle((legend_x, y, legend_x + 18, y + 18), fill=color)
            draw.text((legend_x + 28, y - 2), f"{label}: {format_hours(value)}", fill=self.color("text_color"), font=font)
        image.save(path)

    def draw_report_line_png(self, path: Path, monthly_line: list[int]) -> None:
        from PIL import Image, ImageDraw, ImageFont

        image = Image.new("RGB", (900, 420), self.color("chart_background_color", "#FFFFFF"))
        draw = ImageDraw.Draw(image)
        try:
            font_title = ImageFont.truetype("arial.ttf", 24)
            font = ImageFont.truetype("arial.ttf", 15)
        except OSError:
            font_title = font = ImageFont.load_default()
        draw.text((28, 24), "Aylık Toplam Mesai", fill=self.color("text_color"), font=font_title)
        max_value = max(monthly_line) if monthly_line else 0
        if max_value <= 0:
            draw.text((390, 200), "Veri yok", fill=self.color("text_color"), font=font)
            image.save(path)
            return
        left, top, right, bottom = 90, 80, 860, 350
        draw.line((left, bottom, right, bottom), fill=self.color("text_color"))
        draw.line((left, top, left, bottom), fill=self.color("text_color"))
        for step in range(5):
            value = max_value * step / 4
            y = bottom - (value / max_value) * (bottom - top)
            draw.line((left - 5, y, right, y), fill=self.color("secondary_color"))
            draw.text((10, y - 8), format_hours(int(value)), fill=self.color("text_color"), font=font)
        points = []
        for index, value in enumerate(monthly_line):
            x = left + index * ((right - left) / 11)
            y = bottom - (value / max_value) * (bottom - top)
            points.append((x, y))
        for first, second in zip(points, points[1:]):
            draw.line((first[0], first[1], second[0], second[1]), fill=self.color("accent_color", "#001F5B"), width=4)
        for index, (x, y) in enumerate(points, start=1):
            draw.ellipse((x - 4, y - 4, x + 4, y + 4), fill=self.color("accent_color", "#001F5B"))
            draw.text((x - 5, bottom + 12), str(index), fill=self.color("text_color"), font=font)
        image.save(path)

    def create_kpi_report_images(self, output_dir: Path, payload: dict) -> dict[str, Path]:
        totals = payload["totals"]
        paths = {
            "loss": output_dir / "iscilik_kaybi.png",
            "extra": output_dir / "ek_kapasite.png",
            "monthly": output_dir / "aylik_mesai.png",
        }
        self.draw_report_pie_png(paths["loss"], "İşçilik Kaybı Dağılımı", [("İzin", totals["izin"], "#0EA5E9"), ("Rapor", totals["rapor"], "#F97316")])
        self.draw_report_pie_png(paths["extra"], "Ek Kapasite Dağılımı", [("Fazla Mesai", totals["fazla_mesai"], "#7C3AED"), ("Resmi Tatil Mesai", totals["resmi"], "#22C55E")])
        self.draw_report_line_png(paths["monthly"], payload["monthly_line"])
        return paths

    def desktop_dir(self) -> Path:
        candidates = [
            Path.home() / "Desktop",
            Path.home() / "OneDrive" / "Desktop",
            Path.home() / "Masaüstü",
            Path.home() / "OneDrive" / "Masaüstü",
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return Path.home()

    def ask_kpi_report_type(self) -> str | None:
        dialog = tk.Toplevel(self)
        dialog.title("KPI Rapor Türü")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=self.color("background_color"))
        selected = tk.StringVar(value="")
        ttk.Label(dialog, text="Oluşturulacak rapor türünü seçin:", font=("Segoe UI", 10, "bold")).pack(padx=18, pady=(16, 10))
        buttons = ttk.Frame(dialog)
        buttons.pack(padx=18, pady=(0, 16), fill="x")

        def choose(value: str) -> None:
            selected.set(value)
            dialog.destroy()

        ttk.Button(buttons, text="PDF", command=lambda: choose("pdf")).pack(side="left", padx=5)
        ttk.Button(buttons, text="Excel", command=lambda: choose("excel")).pack(side="left", padx=5)
        ttk.Button(buttons, text="PDF + Excel", command=lambda: choose("both")).pack(side="left", padx=5)
        ttk.Button(buttons, text="İptal", command=dialog.destroy).pack(side="left", padx=5)
        dialog.update_idletasks()
        x = self.winfo_rootx() + max(0, (self.winfo_width() - dialog.winfo_width()) // 2)
        y = self.winfo_rooty() + max(0, (self.winfo_height() - dialog.winfo_height()) // 2)
        dialog.geometry(f"+{x}+{y}")
        self.wait_window(dialog)
        return selected.get() or None

    def create_pdf_report(self, path: Path, person: str, month: int, year: int, cards: list[tuple[str, str]], payload: dict, images: dict[str, Path]) -> None:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import Image as PdfImage
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        font_name = "Helvetica"
        font_bold = "Helvetica-Bold"
        arial_path = Path("C:/Windows/Fonts/arial.ttf")
        arial_bold_path = Path("C:/Windows/Fonts/arialbd.ttf")
        if arial_path.exists():
            pdfmetrics.registerFont(TTFont("ArialTR", str(arial_path)))
            font_name = "ArialTR"
        if arial_bold_path.exists():
            pdfmetrics.registerFont(TTFont("ArialTR-Bold", str(arial_bold_path)))
            font_bold = "ArialTR-Bold"
        doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=1.2 * cm, leftMargin=1.2 * cm, topMargin=1.1 * cm, bottomMargin=1.1 * cm)
        styles = getSampleStyleSheet()
        for style in styles.byName.values():
            style.fontName = font_name
        styles["Title"].fontName = font_bold
        styles["Heading2"].fontName = font_bold
        story = [
            Paragraph("HS Personel Sistemi", styles["Title"]),
            Paragraph("Personel KPI Raporu", styles["Heading2"]),
            Paragraph(f"Personel: {person} &nbsp;&nbsp; Ay: {month} &nbsp;&nbsp; Yıl: {year} &nbsp;&nbsp; Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles["Normal"]),
            Spacer(1, 10),
        ]
        card_table = Table([["KPI", "Değer"]] + cards, colWidths=[9 * cm, 7 * cm])
        card_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8B4B8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8A0A5")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF1F2")]),
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            ("FONTNAME", (0, 1), (-1, -1), font_name),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story += [card_table, Spacer(1, 12)]
        story += [PdfImage(str(images["loss"]), width=8.2 * cm, height=4.5 * cm), Spacer(1, 8)]
        story += [PdfImage(str(images["extra"]), width=8.2 * cm, height=4.5 * cm), PageBreak()]
        story += [Paragraph("Aylık Toplam Mesai", styles["Heading2"]), PdfImage(str(images["monthly"]), width=17 * cm, height=7.9 * cm), Spacer(1, 12)]
        top_rows = [["Personel", "Fazla Mesai", "Resmi Tatil Mesai", "Toplam Mesai", "Limit Kullanım %"]]
        for row in payload.get("top_rows", [])[:25]:
            top_rows.append([row["person"], format_hours(row["fazla_mesai"]), format_hours(row["resmi"]), format_hours(row["toplam"]), f"%{row['limit_oran']:.2f}"])
        top_table = Table(top_rows, repeatRows=1)
        top_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#001F5B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF1F2")]),
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            ("FONTNAME", (0, 1), (-1, -1), font_name),
        ]))
        story += [Paragraph("Top Mesai Listesi", styles["Heading2"]), top_table]
        doc.build(story)

    def create_excel_report(self, path: Path, person: str, month: int, year: int, cards: list[tuple[str, str]], payload: dict, images: dict[str, Path]) -> None:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XlsxImage
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "KPI Özeti"
        ws.append(["HS Personel Sistemi - Personel KPI Raporu"])
        ws.append(["Personel", person, "Ay", month, "Yıl", year, "Oluşturulma", datetime.now().strftime("%d.%m.%Y %H:%M")])
        ws.append([])
        ws.append(["KPI", "Değer"])
        for label, value in cards:
            ws.append([label, value])
        for cell in ws[1]:
            cell.font = Font(bold=True, size=14)
        for cell in ws[4]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E8B4B8")
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 22

        top = wb.create_sheet("Top Mesai Listesi")
        top.append(["Personel", "Fazla Mesai", "Resmi Tatil Mesai", "Toplam Mesai", "Limit Kullanım %"])
        for row in payload.get("top_rows", []):
            top.append([row["person"], format_hours(row["fazla_mesai"]), format_hours(row["resmi"]), format_hours(row["toplam"]), row["limit_oran"] / 100])
        for cell in top[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="001F5B")
        for row in top.iter_rows(min_row=2, min_col=5, max_col=5):
            row[0].number_format = "0.00%"
        for column in "ABCDE":
            top.column_dimensions[column].width = 22

        raw = wb.create_sheet("Ham Veriler")
        raw_rows = self.kpi_raw_rows_for_report(person, month, year)
        raw.append(self.personnel_record_columns)
        for row in raw_rows:
            raw.append([row.get(column, "") for column in self.personnel_record_columns])
        for cell in raw[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E8B4B8")
        for column in raw.columns:
            raw.column_dimensions[column[0].column_letter].width = 18

        charts = wb.create_sheet("Grafikler")
        charts["A1"] = "İşçilik Kaybı Dağılımı"
        charts["J1"] = "Ek Kapasite Dağılımı"
        charts["A24"] = "Aylık Toplam Mesai"
        charts.add_image(XlsxImage(str(images["loss"])), "A2")
        charts.add_image(XlsxImage(str(images["extra"])), "J2")
        charts.add_image(XlsxImage(str(images["monthly"])), "A25")
        wb.save(path)

    def create_html_report(self, path: Path, person: str, month: int, year: int, cards: list[tuple[str, str]], payload: dict, images: dict[str, Path]) -> None:
        def img_data(image_path: Path) -> str:
            return base64.b64encode(image_path.read_bytes()).decode("ascii")

        card_html = "\n".join(f"<div class='card'><span>{html.escape(label)}</span><strong>{html.escape(value)}</strong></div>" for label, value in cards)
        top_html = "\n".join(
            "<tr>"
            f"<td>{html.escape(row['person'])}</td><td>{format_hours(row['fazla_mesai'])}</td><td>{format_hours(row['resmi'])}</td>"
            f"<td>{format_hours(row['toplam'])}</td><td>%{row['limit_oran']:.2f}</td>"
            "</tr>"
            for row in payload.get("top_rows", [])
        )
        content = f"""<!doctype html>
<html lang="tr"><head><meta charset="utf-8"><title>KPI Raporu</title>
<style>
body{{font-family:Segoe UI,Arial,sans-serif;background:#FFF1F2;color:#2B2B2B;margin:0;padding:28px}}
.shell{{max-width:1180px;margin:auto}} h1{{margin:0;color:#001F5B}} h2{{margin-top:4px}}
.meta{{margin:12px 0 22px;color:#555}} .grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px}}
.card{{background:#FFFDFD;border:1px solid #F6D6D8;border-radius:8px;padding:14px;border-top:4px solid #001F5B}}
.card span{{display:block;font-size:13px}} .card strong{{display:block;margin-top:8px;font-size:22px;color:#001F5B}}
.charts{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:18px}} .panel{{background:white;border-radius:8px;padding:12px;border:1px solid #F6D6D8}}
.panel img{{max-width:100%;display:block}} table{{width:100%;border-collapse:collapse;background:white;margin-top:18px}}
th{{background:#001F5B;color:white}} td,th{{padding:9px;border:1px solid #eee;text-align:left}} tr:nth-child(even){{background:#FFF1F2}}
</style></head><body><main class="shell">
<h1>HS Personel Sistemi</h1><h2>Personel KPI Raporu</h2>
<div class="meta">Personel: {html.escape(person)} | Ay: {month} | Yıl: {year} | Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}</div>
<section class="grid">{card_html}</section>
<section class="charts"><div class="panel"><img src="data:image/png;base64,{img_data(images['loss'])}"></div><div class="panel"><img src="data:image/png;base64,{img_data(images['extra'])}"></div></section>
<section class="panel" style="margin-top:16px"><img src="data:image/png;base64,{img_data(images['monthly'])}"></section>
<h2>Top Mesai Listesi</h2><table><thead><tr><th>Personel</th><th>Fazla Mesai</th><th>Resmi Tatil Mesai</th><th>Toplam Mesai</th><th>Limit Kullanım %</th></tr></thead><tbody>{top_html}</tbody></table>
</main></body></html>"""
        path.write_text(content, encoding="utf-8")

    def create_kpi_report(self) -> None:
        report_type = self.ask_kpi_report_type()
        if not report_type:
            return
        if not self.ensure_report_modules_available():
            return
        try:
            person, month, year, _person_count, payload = self.current_kpi_report_payload()
        except Exception as exc:
            messagebox.showerror("KPI Raporu", f"Rapor verisi hazırlanamadı: {exc}")
            return
        try:
            output_dir = self.desktop_dir()
            images = self.create_kpi_report_images(output_dir, payload)
            cards = self.kpi_report_cards(payload["totals"])
            base_name = f"KPI_Raporu_{year}_{month:02d}"
            pdf_path = output_dir / f"{base_name}.pdf"
            xlsx_path = output_dir / f"{base_name}.xlsx"
            created = []
            if report_type in {"pdf", "both"}:
                self.create_pdf_report(pdf_path, person, month, year, cards, payload, images)
                created.append(str(pdf_path))
            if report_type in {"excel", "both"}:
                self.create_excel_report(xlsx_path, person, month, year, cards, payload, images)
                created.append(str(xlsx_path))
        except ModuleNotFoundError as exc:
            missing_name = exc.name or "bilinmiyor"
            messagebox.showerror("KPI Raporu", self.report_dependency_message([missing_name]))
            return
        except Exception as exc:
            messagebox.showerror("KPI Raporu", f"Rapor oluşturulamadı: {exc}")
            return
        messagebox.showinfo("KPI Raporu", "Rapor oluşturuldu ve Masaüstüne kaydedildi:\n" + "\n".join(created))

    def clear_canvas(self, canvas: tk.Canvas, title: str) -> tuple[int, int]:
        canvas.delete("all")
        width = max(canvas.winfo_width(), 420)
        height = max(canvas.winfo_height(), 240)
        canvas.create_text(12, 12, text=title, anchor="nw", fill=self.color("text_color"), font=("", 11, "bold"))
        return width, height

    def draw_pie_chart(self, totals: dict[str, int]) -> None:
        width, _height = self.clear_canvas(self.pie_canvas, "Türlere Göre Saat Dağılımı")
        values = [
            ("Fazla Mesai", totals["fazla_mesai"], "#7C3AED"),
            ("İzin", totals["izin"], "#0EA5E9"),
            ("Rapor", totals["rapor"], "#F97316"),
            ("Resmi Tatil", totals["resmi"], "#22C55E"),
        ]
        total = sum(value for _label, value, _color in values)
        if total <= 0:
            self.pie_canvas.create_text(width / 2, 130, text="Veri yok", fill=self.color("text_color"))
            return
        x0, y0, x1, y1 = 40, 50, 210, 220
        start_angle = 0
        for label, value, color in values:
            extent = value / total * 360
            self.pie_canvas.create_arc(x0, y0, x1, y1, start=start_angle, extent=extent, fill=color, outline=self.color("table_background_color"))
            if value > 0:
                angle = start_angle + extent / 2
                radians = angle * 3.141592653589793 / 180
                label_x = 125 + 56 * math.cos(radians)
                label_y = 135 - 56 * math.sin(radians)
                self.pie_canvas.create_text(
                    label_x,
                    label_y,
                    text=f"%{value / total * 100:.1f}",
                    fill="#FFFFFF",
                    font=("", 8, "bold"),
                )
            start_angle += extent
        legend_x = 240
        for index, (label, value, color) in enumerate(values):
            y = 60 + index * 28
            self.pie_canvas.create_rectangle(legend_x, y, legend_x + 14, y + 14, fill=color, outline=color)
            self.pie_canvas.create_text(legend_x + 20, y + 7, text=f"{label}: {format_hours(value)}", anchor="w", fill=self.color("text_color"))

    def draw_bar_chart(self, by_day: dict[int, dict[str, int]]) -> None:
        width, height = self.clear_canvas(self.bar_canvas, "Gün Bazlı İzin / Rapor / Fazla Mesai")
        if not by_day:
            self.bar_canvas.create_text(width / 2, 130, text="Veri yok", fill=self.color("text_color"))
            return
        days = sorted(by_day)
        max_value = max(
            values.get("izin", 0) + values.get("rapor", 0) + values.get("mesai", 0) + values.get("resmi", 0)
            for values in by_day.values()
        ) or 1
        total_value = sum(
            values.get("izin", 0) + values.get("rapor", 0) + values.get("mesai", 0) + values.get("resmi", 0)
            for values in by_day.values()
        )
        left, top, bottom = 36, 42, height - 28
        chart_width = width - left - 20
        bar_width = max(4, chart_width / max(len(days), 1) * 0.6)
        colors = {"mesai": "#7C3AED", "izin": "#0EA5E9", "rapor": "#F97316", "resmi": "#22C55E"}
        for index, day in enumerate(days):
            x = left + index * (chart_width / max(len(days), 1)) + 4
            y_cursor = bottom
            day_total = 0
            for key in ["mesai", "izin", "rapor", "resmi"]:
                value = by_day[day].get(key, 0)
                day_total += value
                h = (value / max_value) * (bottom - top)
                self.bar_canvas.create_rectangle(x, y_cursor - h, x + bar_width, y_cursor, fill=colors[key], outline="")
                y_cursor -= h
            if day_total > 0:
                percent = day_total / total_value * 100 if total_value > 0 else 0
                self.bar_canvas.create_text(
                    x + bar_width / 2,
                    max(top + 10, y_cursor - 12),
                    text=f"{format_hours(day_total)}\n%{percent:.1f}",
                    fill=self.color("text_color"),
                    font=("", 7),
                    justify="center",
                )
            if index % max(1, len(days) // 10) == 0:
                self.bar_canvas.create_text(x + bar_width / 2, bottom + 10, text=str(day), fill=self.color("text_color"), font=("", 8))

    def draw_line_chart(self, monthly_line: list[int]) -> None:
        width, height = self.clear_canvas(self.line_canvas, "Aylık Toplam Mesai")
        max_value = max(monthly_line) if monthly_line else 0
        if max_value <= 0:
            self.line_canvas.create_text(width / 2, 130, text="Veri yok", fill=self.color("text_color"))
            return
        left, right, top, bottom = 72, width - 28, 46, height - 38
        points = []
        for index, value in enumerate(monthly_line):
            x = left + index * ((right - left) / 11)
            y = bottom - (value / max_value) * (bottom - top)
            points.append((x, y))
        self.line_canvas.create_line(left, bottom, right, bottom, fill=self.color("text_color"))
        self.line_canvas.create_line(left, top, left, bottom, fill=self.color("text_color"))
        for step in range(5):
            value = max_value * step / 4
            y = bottom - (value / max_value) * (bottom - top)
            self.line_canvas.create_line(left - 4, y, right, y, fill=self.color("secondary_color"))
            self.line_canvas.create_text(left - 8, y, text=format_hours(int(value)), anchor="e", fill=self.color("text_color"), font=("", 8))
        for first, second in zip(points, points[1:]):
            self.line_canvas.create_line(first[0], first[1], second[0], second[1], fill=self.color("primary_color"), width=3)
        for index, (x, y) in enumerate(points, start=1):
            self.line_canvas.create_oval(x - 3, y - 3, x + 3, y + 3, fill=self.color("primary_color"), outline="")
            self.line_canvas.create_text(x, bottom + 12, text=str(index), fill=self.color("text_color"), font=("", 8))

    def draw_two_part_pie(self, canvas: tk.Canvas, title: str, values: list[tuple[str, int, str]]) -> None:
        width, _height = self.clear_canvas(canvas, title)
        total = sum(value for _label, value, _color in values)
        if total <= 0:
            canvas.create_text(width / 2, 130, text="Veri yok", fill=self.color("text_color"))
            return
        x0, y0, x1, y1 = 40, 50, 210, 220
        start_angle = 0
        for _label, value, color in values:
            extent = value / total * 360
            canvas.create_arc(x0, y0, x1, y1, start=start_angle, extent=extent, fill=color, outline=self.color("chart_background_color", "#FFFFFF"))
            if value > 0:
                angle = start_angle + extent / 2
                radians = angle * 3.141592653589793 / 180
                label_x = 125 + 56 * math.cos(radians)
                label_y = 135 - 56 * math.sin(radians)
                canvas.create_text(
                    label_x,
                    label_y,
                    text=f"%{value / total * 100:.1f}",
                    fill="#FFFFFF",
                    font=("", 8, "bold"),
                )
            start_angle += extent
        legend_x = 240
        for index, (label, value, color) in enumerate(values):
            y = 60 + index * 28
            canvas.create_rectangle(legend_x, y, legend_x + 14, y + 14, fill=color, outline=color)
            canvas.create_text(legend_x + 20, y + 7, text=f"{label}: {format_hours(value)}", anchor="w", fill=self.color("text_color"))

    def draw_loss_chart(self, totals: dict[str, int]) -> None:
        self.draw_two_part_pie(
            self.loss_canvas,
            "İşçilik Kaybı Dağılımı",
            [
                ("İzin", totals["izin"], "#0EA5E9"),
                ("Rapor", totals["rapor"], "#F97316"),
            ],
        )

    def draw_extra_capacity_chart(self, totals: dict[str, int]) -> None:
        self.draw_two_part_pie(
            self.extra_capacity_canvas,
            "Ek Kapasite Dağılımı",
            [
                ("Fazla Mesai", totals["fazla_mesai"], "#7C3AED"),
                ("Resmi Tatil Mesai", totals["resmi"], "#22C55E"),
            ],
        )

    def refresh_table(self) -> None:
        active_filters = self.active_filters()
        if self.filtering_rows:
            return
        if active_filters and len(self.rows) > self.page_size * 20 and not self.filtering_rows:
            self.start_filter_worker(active_filters)
            return
        self.filtered_indices = self.build_filtered_indices(active_filters)
        self.render_current_page()

    def render_current_page(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)

        total_pages = max(1, (len(self.filtered_indices) + self.page_size - 1) // self.page_size)
        if self.current_page >= total_pages:
            self.current_page = total_pages - 1
        if self.current_page < 0:
            self.current_page = 0
        start_index = self.current_page * self.page_size
        page_indices = self.filtered_indices[start_index:start_index + self.page_size]

        visible_rows: list[dict[str, str]] = []
        for display_index, index in enumerate(page_indices):
            row = self.rows[index]
            try:
                display = self.display_row(row)
            except Exception as exc:
                display = {column: row.get(column, "") for column in self.columns}
                if self.columns:
                    display[self.columns[-1]] = str(exc)
            visible_rows.append(row)
            tag = "even" if display_index % 2 else "odd"
            self.tree.insert("", "end", iid=str(index), values=[display.get(column, "") for column in self.columns], tags=(tag,))

        totals = calculate_totals(visible_rows, self.rules) if visible_rows else {
            "izin_dk": 0,
            "mesai_dk": 0,
            "rapor_dk": 0,
            "resmi_tatil_mesai_dk": 0,
        }
        self.summary_vars["izin"].set(format_hours(totals["izin_dk"]))
        self.summary_vars["mesai"].set(format_hours(totals["mesai_dk"]))
        self.summary_vars["rapor"].set(format_hours(totals["rapor_dk"]))
        self.summary_vars["resmi"].set(format_hours(totals["resmi_tatil_mesai_dk"]))
        shown_start = 0 if not self.filtered_indices else start_index + 1
        shown_end = min(start_index + len(page_indices), len(self.filtered_indices))
        self.page_var.set(f"Sayfa {self.current_page + 1}/{total_pages} | {shown_start}-{shown_end} / {len(self.filtered_indices)}")

    def active_filters(self) -> dict[str, str]:
        return {
            column: variable.get().strip().casefold()
            for column, variable in self.filter_vars.items()
            if variable.get().strip()
        }

    def build_filtered_indices(self, active_filters: dict[str, str] | None = None) -> list[int]:
        active_filters = active_filters or self.active_filters()
        if not active_filters:
            return list(range(len(self.rows)))

        calculated_filter_columns = set(CALCULATED_KEYS).intersection(active_filters)
        result: list[int] = []
        for index, row in enumerate(self.rows):
            if calculated_filter_columns:
                try:
                    searchable = self.display_row(row)
                except Exception:
                    searchable = {column: row.get(column, "") for column in self.columns}
            else:
                searchable = row
            if self.row_matches_filters(searchable, active_filters):
                result.append(index)
        return result

    def start_filter_worker(self, active_filters: dict[str, str]) -> None:
        self.filter_generation += 1
        generation = self.filter_generation
        self.filtering_rows = True
        self.status_var.set("Filtre uygulanıyor...")
        self.progress.start(10)
        thread = threading.Thread(target=self.filter_worker, args=(generation, dict(active_filters)), daemon=True)
        thread.start()
        self.after(100, self.poll_csv_queue)

    def filter_worker(self, generation: int, active_filters: dict[str, str]) -> None:
        try:
            calculated_filter_columns = set(CALCULATED_KEYS).intersection(active_filters)
            result: list[int] = []
            for index, row in enumerate(self.rows):
                if calculated_filter_columns:
                    try:
                        searchable = self.display_row(row)
                    except Exception:
                        searchable = {column: row.get(column, "") for column in self.columns}
                else:
                    searchable = row
                if self.row_matches_filters(searchable, active_filters):
                    result.append(index)
            self.csv_queue.put(("filter_done", (generation, result)))
        except Exception as exc:
            self.csv_queue.put(("filter_error", str(exc)))

    def row_matches_filters(self, display: dict[str, str], active_filters: dict[str, str] | None = None) -> bool:
        filters = active_filters or {
            column: variable.get().strip().casefold()
            for column, variable in self.filter_vars.items()
            if variable.get().strip()
        }
        for column, needle in filters.items():
            if column == "İsim" and normalize_search(needle) == normalize_search("Herkes"):
                continue
            if needle and needle not in str(display.get(column, "")).casefold():
                return False
        return True

    def previous_page(self) -> None:
        if self.current_page > 0:
            self.current_page -= 1
            self.refresh_table()

    def next_page(self) -> None:
        total_pages = max(1, (len(self.filtered_indices) + self.page_size - 1) // self.page_size)
        if self.current_page + 1 < total_pages:
            self.current_page += 1
            self.refresh_table()

    def clear_filters(self) -> None:
        for variable in self.filter_vars.values():
            variable.set("")

    def on_filter_changed(self) -> None:
        self.current_page = 0
        if self.filtering_rows:
            self.filter_generation += 1
            self.filtering_rows = False
            self.progress.stop()
        self.refresh_table()

    def load_selected_into_form(self, _event=None) -> None:
        selected = self.tree.selection()
        if selected:
            index = int(selected[0])
            if 0 <= index < len(self.rows):
                self.fill_form(self.rows[index])

    def copy_selected_rows(self, _event=None):
        selected = self.tree.selection()
        if not selected:
            return "break"
        lines = ["\t".join(self.personnel_record_columns)]
        for item in selected:
            index = int(item)
            if 0 <= index < len(self.rows):
                lines.append("\t".join(self.personnel_export_values(self.rows[index])))
        self.clipboard_clear()
        self.clipboard_append("\n".join(lines))
        return "break"

    def paste_rows(self, _event=None):
        try:
            text = self.clipboard_get()
        except tk.TclError:
            return "break"

        added = 0
        for raw_line in text.splitlines():
            if not raw_line.strip():
                continue
            parts = raw_line.split("\t")
            if len(parts) == 1:
                parts = next(csv.reader([raw_line]))
            parts = [part.strip() for part in parts]
            if [part.casefold() for part in parts[: len(self.input_columns)]] == [column.casefold() for column in self.input_columns]:
                continue
            if len(parts) < len(self.input_columns):
                continue
            row = dict(zip(self.input_columns, parts[: len(self.input_columns)]))
            try:
                calculate_row(row, self.rules)
            except Exception:
                continue
            self.rows.append(row)
            added += 1
        if added:
            self.refresh_kpi_person_options()
            self.refresh_table()
        return "break"

    def default_times_for_date(self, day: date) -> tuple[str, str]:
        if self.rules.full_holiday_on(day) or self.rules.half_holidays_on(day):
            return self.config_data["work_hours"]["weekday_start"], self.config_data["work_hours"]["weekday_end"]
        if day.weekday() == 5 and self.config_data["work_hours"].get("saturday_enabled", True):
            return self.config_data["work_hours"]["saturday_start"], self.config_data["work_hours"]["saturday_end"]
        if day.weekday() == 6 and self.config_data["work_hours"].get("sunday_enabled", False):
            return self.config_data["work_hours"]["sunday_start"], self.config_data["work_hours"]["sunday_end"]
        return self.config_data["work_hours"]["weekday_start"], self.config_data["work_hours"]["weekday_end"]

    def fill_example(self) -> None:
        today = date.today()
        start_time, end_time = self.default_times_for_date(today)
        defaults = self.config_data.get("defaults", {})
        row = {
            "İsim": defaults.get("example_name", defaults.get("İsim", "")),
            "Departman": defaults.get("example_department", defaults.get("Departman", "")),
            "Tür": defaults.get("example_type", defaults.get("Tür", "")),
            "Başlangıç Tarihi": self.format_date(today),
            "Başlangıç Saati": start_time,
            "Bitiş Tarihi": self.format_date(today),
            "Bitiş Saati": end_time,
            "Açıklama": defaults.get("example_description", defaults.get("Açıklama", "")),
        }
        self.fill_form(row)

    def open_csv(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("CSV dosyaları", "*.csv"), ("Tüm dosyalar", "*.*")])
        if not path:
            return
        if self.loading_csv:
            messagebox.showwarning("CSV", "Bir CSV dosyası zaten yükleniyor.")
            return
        self.loading_csv = True
        self.rows = []
        self.current_page = 0
        self.status_var.set("CSV yükleniyor...")
        self.progress.start(10)
        thread = threading.Thread(target=self.load_csv_worker, args=(path,), daemon=True)
        thread.start()
        self.after(100, self.poll_csv_queue)

    def load_csv_worker(self, path: str) -> None:
        loaded_rows: list[dict[str, str]] = []
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as file:
                reader = csv.DictReader(file)
                for count, row in enumerate(reader, start=1):
                    loaded_rows.append({column: (row.get(column) or "").strip() for column in self.input_columns})
                    if count % self.csv_progress_interval == 0:
                        self.csv_queue.put(("progress", count))
            self.csv_queue.put(("done", loaded_rows))
        except Exception as exc:
            self.csv_queue.put(("error", str(exc)))

    def poll_csv_queue(self) -> None:
        try:
            while True:
                message_type, payload = self.csv_queue.get_nowait()
                if message_type == "progress":
                    self.status_var.set(f"CSV yükleniyor... {payload:,} satır okundu")
                elif message_type == "done":
                    self.rows = payload
                    self.loading_csv = False
                    self.progress.stop()
                    self.status_var.set(f"CSV yüklendi: {len(self.rows):,} satır")
                    self.current_page = 0
                    self.refresh_kpi_person_options()
                    self.refresh_table()
                elif message_type == "error":
                    self.loading_csv = False
                    self.progress.stop()
                    self.status_var.set("CSV yüklenemedi")
                    messagebox.showerror("CSV açılamadı", payload)
                elif message_type == "filter_done":
                    generation, indices = payload
                    if generation == self.filter_generation:
                        self.filtered_indices = indices
                        self.filtering_rows = False
                        self.progress.stop()
                        self.status_var.set(f"Filtre tamamlandı: {len(indices):,} satır")
                        self.current_page = 0
                        self.render_current_page()
                elif message_type == "filter_error":
                    self.filtering_rows = False
                    self.progress.stop()
                    self.status_var.set("Filtre uygulanamadı")
                    messagebox.showerror("Filtre hatası", payload)
        except queue.Empty:
            pass
        if self.loading_csv or self.filtering_rows:
            self.after(100, self.poll_csv_queue)

    def save_csv(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV dosyaları", "*.csv"), ("Tüm dosyalar", "*.*")])
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as file:
                writer = csv.DictWriter(file, fieldnames=self.personnel_record_columns)
                writer.writeheader()
                for row in self.rows:
                    writer.writerow(self.personnel_export_row(row))
        except Exception as exc:
            messagebox.showerror("CSV kaydedilemedi", str(exc))

    def save_excel(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel dosyalarÄ±", "*.xlsx"), ("TÃ¼m dosyalar", "*.*")])
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ModuleNotFoundError as exc:
            missing = exc.name or "openpyxl"
            messagebox.showerror("Excel kaydedilemedi", f"Excel dÄ±ÅŸa aktarma iÃ§in gerekli kÃ¼tÃ¼phane eksik.\n\nEksik modÃ¼l: {missing}\n\nKurulum:\npip install openpyxl")
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Personel KayÄ±tlarÄ±"
            ws.append(self.personnel_record_columns)
            for row in self.rows:
                ws.append(self.personnel_export_values(row))
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E8B4B8")
            for column_cells in ws.columns:
                column_letter = column_cells[0].column_letter
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 32)
            wb.save(path)
        except Exception as exc:
            messagebox.showerror("Excel kaydedilemedi", str(exc))

def main() -> None:
    try:
        config = load_config()
    except ConfigError as exc:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Config Hatası", str(exc))
        return
    app = PersonelTakipApp(config)
    app.mainloop()


if __name__ == "__main__":
    main()
